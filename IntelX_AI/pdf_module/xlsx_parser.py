"""
XLSX/Excel file parser module.
Extracts data from Excel files and converts to text format for LLM processing.
"""
from pathlib import Path
from typing import Dict, List, Any, Optional
import sys
import os

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from exceptions import XLSXParserError, FileHandlerError, DependencyError


def parse_xlsx_file(xlsx_path: str) -> str:
    """
    Parse XLSX file and convert to text format for LLM extraction.
    
    Args:
        xlsx_path: Path to the XLSX file
        
    Returns:
        String representation of Excel data
    """
    if not xlsx_path:
        raise XLSXParserError("XLSX file path is required", xlsx_path)
    
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise DependencyError(
            "openpyxl is required to read XLSX files. Install it with `pip install openpyxl`.",
            xlsx_path
        )
    
    path = Path(xlsx_path)
    
    if not path.exists():
        raise FileHandlerError(f"XLSX file not found: {xlsx_path}", xlsx_path)
    
    if not path.is_file():
        raise FileHandlerError(f"Path is not a file: {xlsx_path}", xlsx_path)
    
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        
        text_parts = []
        text_parts.append(f"Excel Workbook: {path.name}")
        text_parts.append(f"Total Sheets: {len(workbook.sheetnames)}")
        text_parts.append("")
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")
            text_parts.append("=" * 50)
            
            # Get all rows
            rows = list(sheet.iter_rows(values_only=True))
            
            if not rows:
                text_parts.append("(Empty sheet)")
                text_parts.append("")
                continue
            
            # First row as headers
            headers = [str(cell) if cell is not None else f"Column_{idx+1}" 
                      for idx, cell in enumerate(rows[0])]
            
            text_parts.append("Headers: " + " | ".join(headers))
            text_parts.append("")
            
            # Data rows
            for row_idx, row in enumerate(rows[1:], start=2):
                text_parts.append(f"Row {row_idx}:")
                for col_idx, cell_value in enumerate(row):
                    if cell_value is not None:
                        header = headers[col_idx] if col_idx < len(headers) else f"Column_{col_idx+1}"
                        text_parts.append(f"  {header}: {str(cell_value)}")
                text_parts.append("")
            
            text_parts.append("")
        
        workbook.close()
        return "\n".join(text_parts)
    
    except FileNotFoundError as e:
        raise FileHandlerError(f"XLSX file not found: {xlsx_path}", xlsx_path) from e
    except PermissionError as e:
        raise FileHandlerError(f"Permission denied reading XLSX file: {xlsx_path}", xlsx_path) from e
    except Exception as e:
        if "openpyxl" in str(type(e).__module__):
            raise XLSXParserError(f"openpyxl error: {str(e)}", xlsx_path) from e
        raise XLSXParserError(f"Unexpected error parsing XLSX file: {str(e)}", xlsx_path) from e


def extract_xlsx_data(xlsx_path: str) -> Dict[str, Any]:
    """
    Extract structured data from XLSX file.
    
    Args:
        xlsx_path: Path to the XLSX file
        
    Returns:
        Dictionary with Excel structure and data
        
    Raises:
        XLSXParserError: If parsing fails
        FileHandlerError: If file operations fail
        DependencyError: If openpyxl is not installed
    """
    if not xlsx_path:
        raise XLSXParserError("XLSX file path is required", xlsx_path)
    
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise DependencyError(
            "openpyxl is required to read XLSX files. Install it with `pip install openpyxl`.",
            xlsx_path
        )
    
    path = Path(xlsx_path)
    
    if not path.exists():
        raise FileHandlerError(f"XLSX file not found: {xlsx_path}", xlsx_path)
    
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        
        sheets_data = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            
            if not rows:
                sheets_data.append({
                    "sheet_name": sheet_name,
                    "headers": [],
                    "data": []
                })
                continue
            
            # First row as headers
            headers = [str(cell) if cell is not None else f"Column_{idx+1}" 
                      for idx, cell in enumerate(rows[0])]
            
            # Convert rows to dictionaries
            data_rows = []
            for row in rows[1:]:
                row_dict = {}
                for col_idx, cell_value in enumerate(row):
                    header = headers[col_idx] if col_idx < len(headers) else f"Column_{col_idx+1}"
                    row_dict[header] = str(cell_value) if cell_value is not None else ""
                data_rows.append(row_dict)
            
            sheets_data.append({
                "sheet_name": sheet_name,
                "headers": headers,
                "data": data_rows,
                "total_rows": len(data_rows)
            })
        
        workbook.close()
        
        return {
            "file_path": str(path),
            "total_sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "sheets": sheets_data
        }
    
    except FileNotFoundError as e:
        raise FileHandlerError(f"XLSX file not found: {xlsx_path}", xlsx_path) from e
    except PermissionError as e:
        raise FileHandlerError(f"Permission denied reading XLSX file: {xlsx_path}", xlsx_path) from e
    except Exception as e:
        if "openpyxl" in str(type(e).__module__):
            raise XLSXParserError(f"openpyxl error: {str(e)}", xlsx_path) from e
        raise XLSXParserError(f"Unexpected error extracting XLSX data: {str(e)}", xlsx_path) from e

